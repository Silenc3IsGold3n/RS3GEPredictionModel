import openpyxl
import re
import sys
import datetime
import json
import requests
import threading
import time
import gatherData
import printData
import getItemIds

path = "Rs3 Investments.xlsx"
wb_obj = openpyxl.load_workbook(path.strip())
# from the active attribute 
sheet_obj = wb_obj.active
max_column=sheet_obj.max_column
max_row=sheet_obj.max_row
items = []
alphabet = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']

def get_Data(x,y):
	lock = threading.Lock()
	start_time = time.time()
	total_items_added = 0
	gatherData.load_urls()
	for i in range(x, y):
		url = 'http://services.runescape.com/m=itemdb_rs/api/catalogue/items.json?category='+ str(i) + '&alpha=a'
		items_in_category = gatherData.get_items_in_category(url)
		print('Items in this category: ' + str(items_in_category))
		current_items = 0
		for j in alphabet:
			url = 'http://services.runescape.com/m=itemdb_rs/api/catalogue/items.json?category='+ str(i) + '&alpha=' + str(j) + '&page=1'
			gatherData.run(url,1,lock)
			current_items = current_items + gatherData.get_current_items()
			print('Items added so far in category: ' + str(current_items) +'/'+str(items_in_category))
			gatherData.reset_current_items()
			if (current_items == items_in_category):
				break
		total_items_added = total_items_added + current_items
	print('Total items added: ' + str(total_items_added))
	print('Toal time: ' + str((time.time() - start_time)) + ' seconds.')

print("grabbing item IDs from spreadsheet")
for j in range(3, max_column+1):#column checking item IDs
    item_id = sheet_obj.cell(row=2,column=j)
    items.append(item_id.value)
items.reverse()
currentrow = 0

print("adding the new date row")
for j in range(14,max_row):#setting date/updating values
    date = sheet_obj.cell(row=j,column=2)
    if date.value is None:
        current_time = datetime.date.today()
        dttm = current_time.strftime('%m/%d/%Y')
        date.value = dttm
        currentrow = j
        break
        
if currentrow == 0:
    currentrow = max_row + 1
print("adding updated prices to new row")
item_url = "https://services.runescape.com/m=itemdb_rs/api/graph/X.json"

amtItems = len(items)
amtSamePrice = 0
for j in range(3, max_column+1):#updating value for gp for today
    price = sheet_obj.cell(row=currentrow,column=j)
    if price.value is None:#use items.pop() for the name to check in database for new GP value
        item = items.pop()
        request = item_url.replace("X", "{}".format(item))
        response = requests.get(request)
        if response.status_code is not 200:
                print("HTTP Error. API Response Was Not 200: " + str(response.status_code))
        else:
            jsonResponse = response.json()
            lastKey = list(jsonResponse['daily'].keys())[-1]
            lastValue = jsonResponse['daily'][lastKey]
            print(lastValue)
            previousValue = sheet_obj.cell(row=currentrow-1,column=j)
            price.value = lastValue
            if(previousValue.value == lastValue or previousValue.value is None): #If amount matches the amount from the previous day, or if it is a newly added item to the sheet
                amtSamePrice += 1
print("amt the same:" + str(amtSamePrice))
print("amt items:" + str(amtItems))
if amtSamePrice == amtItems:
    print("GE isnt updated yet, all values the same. Task should be scheduled to restart this 4 hours from now.")
    wb_obj.close()
    exit(1)
wb_obj.save(path)
wb_obj.close()


#Update Database:
#for i in range(0,1): 
    #lower = 0 + (i*18)
    #upper = (18+i) + (i*18)
lower = 0
upper = 37
mythread = threading.Thread(name = "Thread-{}".format(1),target = get_Data,kwargs={'x': lower,'y': upper}) 
mythread.start()
time.sleep(.1)
#get_Data(0,37)